"""
==========================================================================
FINAL RESULTS BUILDER
==========================================================================

This script produces the final results for accuracy, row-level Actions ROC-AUC,
inter-rater reliability, and response-change analyses across the four labeling
tasks: Activity Type, Actions Occurring, Engagement, and Gaze.

Outputs:
    Final_Result_Clean.xlsx
        A complete analysis workbook containing normalized ground truth,
        participant-level response scoring, Actions ROC-AUC detail and summaries,
        modality and participant summaries,
        validation records, Engagement and Gaze scoring, Within-subject change
        details and summaries, Engagement Delta audit records, reliability
        documentation, 43 Between-subject
        reliability_data matrix sheets, matrix indexes, option-level Alpha
        results, and Alpha summaries.

    Final_Table.xlsx
        Four paper-ready summary tables covering accuracy, inter-rater
        reliability, response-change rates, and engagement-change direction.
"""

from __future__ import annotations

from collections import defaultdict
from copy import copy
import math
from pathlib import Path
import re
import tempfile

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ========================================================================
# PART 1. FILE PATHS AND FIXED STUDY DEFINITIONS
# ========================================================================

# Inputs, outputs, and optional local dependencies live beside this script.
ROOT_DIR = Path(__file__).resolve().parent
SURVEY_INPUT = ROOT_DIR / "Qualtrics_VideoAnonymization_PilotStudy.xlsx"
GROUND_TRUTH_INPUT = ROOT_DIR / "Ground_truth_labels.xlsx"
FINAL_RESULT_CLEAN = ROOT_DIR / "Final_Result_Clean.xlsx"
FINAL_TABLE = ROOT_DIR / "Final_Table.xlsx"

MODALITIES = [
    "Rendered without sound",
    "Rendered with sound",
    "Blurred",
    "Raw",
]
DISPLAY_MODALITY = {
    "Rendered without sound": "Rendered (no sound)",
    "Rendered with sound": "Rendered",
    "Rendered": "Rendered",
    "Blurred": "Blurred",
    "Raw": "Raw",
}
TASKS = ["Activity Type", "Actions Occurring", "Engagement", "Gaze"]
ENGAGEMENT_SCALE = {"Low": 0, "Medium": 1, "High": 2}
ACTIVITY_CODES = {
    "individual activity": 0.0,
    "small group activity": 1.0,
    "whole class activity": 2.0,
}
ACTION_OPTIONS = [
    "is a student moving around",
    "is a student raising hand",
    "is a student speaking",
    "is a student writing or working",
    "is a teacher sitting",
    "is a teacher speaking",
    "is a teacher standing",
    "is a teacher walking",
]
ENGAGEMENT_CODES = {"Low": 0.0, "Medium": 1.0, "High": 2.0}
GAZE_CODES = {"ID 1": 0.0, "ID 2": 1.0, "ID 3": 2.0, "ID 4": 3.0}

# Fixed survey-instrument schema. The raw Qualtrics export contains technical
# field names but not the human-readable question-text row. These constants
# preserve the validated field-to-stimulus mapping without requiring a third
# input document.
BETWEEN_ACTIVITY_CLIPS = [
    "video12", "video2", "video3", "video4", "video5", "video6",
    "video7", "video8", "video9", "VsIfRjWJRuk_00-14-15_to_00-14-25",
    "weoneSV3lnU_00-01-15_to_00-01-25",
    "weoneSV3lnU_00-02-22_to_00-02-32",
    "WpvHi9iiaKk_00-01-00_to_00-01-10",
    "Z9tJBluH1Zw_00-00-13_to_00-00-23",
    "Z9tJBluH1Zw_00-14-58_to_00-15-08",
]
BETWEEN_ENGAGEMENT_CLIPS = [
    "view2194", "view2195", "view2196", "view2197", "view2003",
    "view2005", "view2007", "view2008", "view12", "view13", "view14",
    "view16", "view2198", "view2010", "view2",
]
WITHIN_ACTIVITY_CLIPS = [
    "0Egr2Xxr95k_00-00-12_to_00-00-22",
    "0Egr2Xxr95k_00-00-20_to_00-00-30",
    "0Egr2Xxr95k_00-01-56_to_00-02-06",
    "0Egr2Xxr95k_00-02-35_to_00-02-45",
]
WITHIN_ENGAGEMENT_CLIPS = [
    "view2191", "view2193", "7", "view2001", "view1", "view11",
]
BETWEEN_GAZE_GROUP_CLIPS = [
    "0Tfgw5y1Xec_00-05-15_to_00-05-25",
    "MNgTBtTz_Qo_00-08-26_to_00-08-36",
    "PZY-hB2C_Iw_00-08-04_to_00-08-14",
    "0Egr2Xxr95k_00-05-34_to_00-05-44",
    "PZY-hB2C_Iw_00-08-36_to_00-08-46",
    "video10",
    "0Egr2Xxr95k_00-07-25_to_00-07-35",
    "0Egr2Xxr95k_00-06-26_to_00-06-36",
    "0Egr2Xxr95k_00-02-35_to_00-02-45",
    "0Egr2Xxr95k_00-08-02_to_00-08-12",
    "PZY-hB2C_Iw_00-08-04_to_00-08-14",
    "MNgTBtTz_Qo_00-05-20_to_00-05-30",
    "0Egr2Xxr95k_00-05-34_to_00-05-44",
    "0Egr2Xxr95k_00-02-35_to_00-02-45",
    "KRgIxK0WNis_00-07-36_to_00-07-46",
]


# ========================================================================
# PART 2. GENERAL READING, VALIDATION, AND FORMATTING HELPERS
# ========================================================================


def require_inputs() -> None:
    """Confirm that the raw survey and ground-truth inputs exist."""
    missing = [
        path for path in (SURVEY_INPUT, GROUND_TRUTH_INPUT)
        if not path.exists()
    ]
    if missing:
        raise FileNotFoundError("Missing required input(s): " + ", ".join(map(str, missing)))


def prepare_analysis_inputs(temporary_directory: Path):
    """
    Build an analysis-ready temporary copy from the single-sheet Qualtrics
    export. Participant labels and fixed question metadata are generated from
    response order and the documented survey schema embedded in this script.
    """
    reconstructed = load_workbook(SURVEY_INPUT)
    responses = reconstructed[reconstructed.sheetnames[0]]
    responses.title = "Sheet 1 - VideoAnonymization_Pi"
    headers = [cell.value for cell in responses[1]]
    response_id_column = headers.index("ResponseId") + 1

    participant_map = reconstructed.create_sheet("Sheet1")
    participant_map.append(["P_label", "ResponseId"])
    participant_number = 0
    for row_number in range(3, responses.max_row + 1):
        response_id = responses.cell(row_number, response_id_column).value
        if not clean(response_id):
            continue
        participant_number += 1
        participant_map.append([f"P{participant_number}", response_id])
    if participant_number != 40:
        raise ValueError(
            f"Expected 40 participant records, found {participant_number}."
        )

    question_text = {
        clean(header): clean(header) for header in headers if clean(header)
    }
    for index, clip in enumerate(BETWEEN_ACTIVITY_CLIPS, start=1):
        question_text[f"{index}_Q17"] = f"{clip} - Which Activity Type"
        question_text[f"{index}_Q13"] = f"{clip} - Which Actions Occurring"
    for index, clip in enumerate(BETWEEN_ENGAGEMENT_CLIPS, start=1):
        question_text[f"{index}_Q33"] = f"{clip} - Engagement"
    for index, clip in enumerate(WITHIN_ACTIVITY_CLIPS, start=1):
        for suffix in ("Q20", "Q21", "Q26", "Q29", "Q25", "Q27", "Q24", "Q28"):
            question_text[f"{index}_{suffix}"] = f"{clip} - Which Within Activity"
    for index, clip in enumerate(WITHIN_ENGAGEMENT_CLIPS, start=1):
        for suffix in ("Q36", "Q37", "Q38", "Q39"):
            question_text[f"{index}_{suffix}"] = f"{clip} - Within Engagement"
    for group, clip in enumerate(BETWEEN_GAZE_GROUP_CLIPS, start=1):
        first_field = (group - 1) * 3 + 1
        for offset in range(3):
            question_text[f"{first_field + offset}_Q53"] = f"{clip} - Gaze"

    metadata = reconstructed.create_sheet("Sheet2")
    metadata.append(headers)
    metadata.append([question_text.get(clean(header), clean(header)) for header in headers])

    temporary_survey = temporary_directory / "survey_input_reconstructed.xlsx"
    reconstructed.save(temporary_survey)
    return temporary_survey, GROUND_TRUTH_INPUT


def rows_as_dicts(worksheet) -> list[dict]:
    """Convert a sheet to list[dict] for Design/Task/Modality filtering."""
    headers = [cell.value for cell in worksheet[1]]
    return [
        dict(zip(headers, row))
        for row in worksheet.iter_rows(min_row=2, values_only=True)
    ]


def clean(value) -> str:
    """Convert None to an empty string and trim all other values."""
    return "" if value is None else str(value).strip()


def safe_number(value):
    """Return a finite Excel-safe number; display NaN/Inf as blank."""
    if value is None:
        return None
    number = float(value)
    return number if math.isfinite(number) else None


NAVY = "17365D"
PALE_BLUE = "DCE6F1"
LIGHT_LINE = Side(style="thin", color="B4C6E7")
BLACK_THIN = Side(style="thin", color="000000")
BLACK_MEDIUM = Side(style="medium", color="000000")


def style_data_header(cells) -> None:
    """Apply the standard dark-blue style to data/index headers."""
    for cell in cells:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=LIGHT_LINE)


def set_widths(worksheet, widths: dict[str, float]) -> None:
    """Set readable column widths by column letter."""
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


# ========================================================================
# PART 3. ANALYSIS PIPELINE
# ========================================================================


def ensure_krippendorff_importable() -> None:
    """
    Make the pinned local krippendorff package available when the script is
    executed with the system Python that provides pandas.
    """
    try:
        __import__("krippendorff")
        return
    except ModuleNotFoundError:
        pass

    import sys

    candidates = sorted((ROOT_DIR / ".venv-krippendorff" / "lib").glob(
        "python*/site-packages"
    ))
    if not candidates:
        raise ModuleNotFoundError(
            "The krippendorff package is unavailable. Install the dependencies "
            "listed in requirements.txt."
        )
    sys.path.append(str(candidates[-1]))
    __import__("krippendorff")


def between_participant_key(value) -> int:
    """Sort participant labels by their numeric suffix (P2 before P10)."""
    match = re.search(r"\d+", clean(value))
    return int(match.group()) if match else 10**9


def between_split_set(value) -> set[str]:
    """Split a normalized comma-separated multi-select response."""
    return {part.strip() for part in clean(value).split(",") if part.strip()}


def prepare_between_matrix_rows(
    rows,
    modality,
    extra_filter=lambda _row: True,
    unit_columns=("Clip_ID",),
):
    """Build the participant/unit axes and response lookup for one modality."""
    between_rows = [
        row for row in rows if row["Design"] == "Between" and extra_filter(row)
    ]
    selected = [row for row in between_rows if row["Modality"] == modality]
    participants = sorted(
        {row["Participant"] for row in between_rows}, key=between_participant_key
    )
    units = sorted(
        {tuple(row[column] for column in unit_columns) for row in between_rows},
        key=lambda values: tuple(clean(value) for value in values),
    )
    lookup = {}
    for row in selected:
        unit = tuple(row[column] for column in unit_columns)
        key = (row["Participant"], unit)
        if key in lookup:
            raise ValueError(f"Duplicate participant × unit: {key}")
        lookup[key] = row
    labels = [" | ".join(clean(value) for value in unit) for unit in units]
    return participants, units, labels, lookup


def between_coded_matrix(
    rows,
    modality,
    value_column,
    codes,
    extra_filter=lambda _row: True,
    unit_columns=("Clip_ID",),
):
    """Create a numeric Between-subject reliability matrix."""
    participants, units, labels, lookup = prepare_between_matrix_rows(
        rows, modality, extra_filter, unit_columns
    )
    matrix = np.full((len(participants), len(units)), np.nan, dtype=float)
    for row_index, participant in enumerate(participants):
        for column_index, unit in enumerate(units):
            row = lookup.get((participant, unit))
            if row is None or not clean(row[value_column]):
                continue
            value = clean(row[value_column])
            if value not in codes:
                raise ValueError(f"Unexpected category {value!r}")
            matrix[row_index, column_index] = codes[value]
    return matrix, participants, labels


def between_action_matrix(rows, modality, option):
    """Create one binary matrix for a single Actions response option."""
    participants, units, labels, lookup = prepare_between_matrix_rows(
        rows,
        modality,
        lambda row: row["Question_Type"] == "Actions",
    )
    matrix = np.full((len(participants), len(units)), np.nan, dtype=float)
    for row_index, participant in enumerate(participants):
        for column_index, unit in enumerate(units):
            row = lookup.get((participant, unit))
            if row is None or not clean(row["Response_Normalized"]):
                continue
            matrix[row_index, column_index] = float(
                option in between_split_set(row["Response_Normalized"])
            )
    return matrix, participants, labels


def between_alpha(matrix, level) -> float:
    """Calculate package-based alpha, preserving single-category results as NaN."""
    import krippendorff

    try:
        return float(
            krippendorff.alpha(
                reliability_data=matrix,
                level_of_measurement=level,
            )
        )
    except ValueError as error:
        if "more than one value in the domain" in str(error):
            return float("nan")
        raise


def calculate_between_reliability(analysis_workbook: Path):
    """Build all 43 corrected Between-subject matrices in memory."""
    source = load_workbook(analysis_workbook, read_only=True, data_only=True)
    try:
        response = rows_as_dicts(source["Response_Scoring"])
        engagement = rows_as_dicts(source["Engagement_Scoring"])
        gaze = rows_as_dicts(source["Gaze_Scoring"])
    finally:
        source.close()

    records = []
    for modality in MODALITIES:
        matrix, participants, units = between_coded_matrix(
            response,
            modality,
            "Response_Normalized",
            ACTIVITY_CODES,
            lambda row: row["Question_Type"] == "Activities",
        )
        records.append((
            "Activity Type", modality, "", matrix, participants, units,
            "nominal", between_alpha(matrix, "nominal"),
        ))

    for modality in MODALITIES:
        for option in ACTION_OPTIONS:
            matrix, participants, units = between_action_matrix(
                response, modality, option
            )
            records.append((
                "Actions Occurring", modality, option, matrix, participants,
                units, "nominal", between_alpha(matrix, "nominal"),
            ))

    for modality in MODALITIES:
        matrix, participants, units = between_coded_matrix(
            engagement, modality, "Response", ENGAGEMENT_CODES
        )
        records.append((
            "Engagement", modality, "", matrix, participants, units,
            "ordinal", between_alpha(matrix, "ordinal"),
        ))

    for modality in ("Rendered with sound", "Blurred", "Raw"):
        source_modality = "Rendered" if modality == "Rendered with sound" else modality
        matrix, participants, units = between_coded_matrix(
            gaze,
            source_modality,
            "Response",
            GAZE_CODES,
            unit_columns=("Clip_ID", "Frame_Idx"),
        )
        records.append((
            "Gaze", modality, "", matrix, participants, units,
            "nominal", between_alpha(matrix, "nominal"),
        ))

    if len(records) != 43:
        raise AssertionError(f"Expected 43 reliability matrices, got {len(records)}")
    return records


def reliability_records_to_content(records):
    """
    Convert the 43 in-memory reliability records into keyed matrices and
    metadata. This is the same information previously stored in an NPZ archive,
    but it now flows directly into the final workbook without serialization.
    """
    matrices = {}
    metadata = {}
    action_counter = defaultdict(int)
    modality_code = {
        "Rendered without sound": "NoSound",
        "Rendered with sound": "Rendered",
        "Blurred": "Blurred",
        "Raw": "Raw",
    }
    task_code = {
        "Activity Type": "Activity",
        "Actions Occurring": "Action",
        "Engagement": "Engage",
        "Gaze": "Gaze",
    }

    for index, record in enumerate(records, start=1):
        task, modality, option, matrix, participants, units, level, alpha = record
        key = f"matrix_{index:02d}"
        if task == "Actions Occurring":
            action_counter[modality] += 1
            sheet = (
                f"{task_code[task]}_{modality_code[modality]}_"
                f"{action_counter[modality]:02d}"
            )
        else:
            sheet = f"{task_code[task]}_{modality_code[modality]}"
        matrices[key] = matrix
        metadata[key] = {
            "sheet": sheet,
            "task": task,
            "modality": DISPLAY_MODALITY[modality],
            "action_option_number": (
                action_counter[modality] if task == "Actions Occurring" else None
            ),
            "action_option": option,
            "measurement_level": level,
            "participants": participants,
            "clip_ids": units,
            "alpha": alpha,
        }

    if len(matrices) != 43:
        raise AssertionError(f"Expected 43 reliability matrices, got {len(matrices)}")
    return matrices, metadata


def add_actions_auc(analysis_workbook: Path) -> None:
    """
    Add row-level Actions ROC-AUC calculations and summary statistics.

    Unit of analysis:
        one participant + one video response row.

    Construction:
        y_true contains one binary value per label in the global Actions label
        universe, indicating whether that label is in the ground truth.
        y_score contains the participant's binary selected/not-selected value
        for the same labels.

    Because y_score is binary rather than a continuous confidence score, the
    resulting ROC-AUC is equivalent to balanced accuracy:

        ROC-AUC = (sensitivity + specificity) / 2

    If y_true has only one class, sklearn cannot define ROC-AUC; that row is
    stored as blank/NaN and excluded from Mean_ROC_AUC.
    """
    from sklearn.metrics import roc_auc_score

    workbook = load_workbook(analysis_workbook)
    worksheet = workbook["Response_Scoring"]
    headers = [cell.value for cell in worksheet[1]]
    header_index = {name: index for index, name in enumerate(headers)}
    required = {
        "Design", "Modality", "Question_Type", "Response_Normalized",
        "Ground_Truth_Normalized", "Answer_Choices",
    }
    missing = required - set(header_index)
    if missing:
        raise ValueError(f"Response_Scoring is missing AUC fields: {sorted(missing)}")

    def label_set(value):
        return {
            item.strip().lower().replace("?", "")
            for item in clean(value).split(",")
            if item.strip()
        }

    raw_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    action_universe = set()
    for row in raw_rows:
        if row[header_index["Question_Type"]] != "Actions":
            continue
        action_universe.update(label_set(row[header_index["Answer_Choices"]]))
        action_universe.update(label_set(row[header_index["Ground_Truth_Normalized"]]))
        action_universe.update(label_set(row[header_index["Response_Normalized"]]))
    action_labels = sorted(action_universe)
    if len(action_labels) < 2:
        raise ValueError("Actions AUC requires at least two labels")

    new_headers = [
        "Action_Universe_Size", "TN", "Specificity", "ROC_AUC", "AUC_Defined"
    ]
    start_column = worksheet.max_column + 1
    for offset, name in enumerate(new_headers):
        worksheet.cell(1, start_column + offset, name)
    style_data_header(worksheet[1])

    detail_rows = []
    for excel_row, row in enumerate(raw_rows, start=2):
        if row[header_index["Question_Type"]] != "Actions":
            continue
        truth = label_set(row[header_index["Ground_Truth_Normalized"]])
        response = label_set(row[header_index["Response_Normalized"]])
        blank_response = not clean(row[header_index["Response_Normalized"]])
        if blank_response:
            tn = specificity = auc = None
            defined = 0
        else:
            y_true = np.asarray([int(label in truth) for label in action_labels])
            y_score = np.asarray([int(label in response) for label in action_labels])
            tn = int(np.sum((y_true == 0) & (y_score == 0)))
            fp = int(np.sum((y_true == 0) & (y_score == 1)))
            specificity = tn / (tn + fp) if tn + fp else None
            defined = int(len(np.unique(y_true)) == 2)
            auc = float(roc_auc_score(y_true, y_score)) if defined else None
        values = [len(action_labels), tn, specificity, auc, defined]
        for offset, value in enumerate(values):
            worksheet.cell(excel_row, start_column + offset, value)
        worksheet.cell(excel_row, start_column + 2).number_format = "0.000000"
        worksheet.cell(excel_row, start_column + 3).number_format = "0.000000"
        detail_rows.append({
            "Design": row[header_index["Design"]],
            "Modality": row[header_index["Modality"]],
            "Specificity": specificity,
            "ROC_AUC": auc,
            "AUC_Defined": defined,
        })

    for column in range(start_column, start_column + len(new_headers)):
        worksheet.column_dimensions[get_column_letter(column)].width = 20

    if "Actions_AUC_Summary" in workbook.sheetnames:
        del workbook["Actions_AUC_Summary"]
    summary = workbook.create_sheet("Actions_AUC_Summary")
    summary.append([
        "Design", "Modality", "Action_Rows", "Valid_AUC_Rows",
        "Undefined_AUC_Rows", "Mean_ROC_AUC", "Median_ROC_AUC",
        "SD_ROC_AUC", "Mean_Specificity",
    ])
    designs = ["All", "Between", "Within"]
    for design in designs:
        for modality in MODALITIES:
            selected = [
                row for row in detail_rows
                if (design == "All" or row["Design"] == design)
                and row["Modality"] == modality
            ]
            valid_auc = [
                row["ROC_AUC"] for row in selected if row["ROC_AUC"] is not None
            ]
            valid_specificity = [
                row["Specificity"]
                for row in selected if row["Specificity"] is not None
            ]
            summary.append([
                design, modality, len(selected), len(valid_auc),
                len(selected) - len(valid_auc),
                mean(valid_auc),
                float(np.median(valid_auc)) if valid_auc else None,
                float(np.std(valid_auc, ddof=1)) if len(valid_auc) > 1 else None,
                mean(valid_specificity),
            ])
    style_data_header(summary[1])
    summary.freeze_panes = "A2"
    summary.sheet_view.showGridLines = False
    set_widths(summary, {
        "A": 14, "B": 28, "C": 15, "D": 18, "E": 22,
        "F": 18, "G": 20, "H": 18, "I": 20,
    })
    for row in summary.iter_rows(min_row=2, min_col=6, max_col=9):
        for cell in row:
            cell.number_format = "0.000000"

    if "Actions_AUC_Method" in workbook.sheetnames:
        del workbook["Actions_AUC_Method"]
    method = workbook.create_sheet("Actions_AUC_Method")
    method_rows = [
        ["Item", "Explanation"],
        ["Package", "sklearn.metrics.roc_auc_score"],
        ["Analysis unit", "One participant + one video Actions response."],
        ["y_true", "Binary ground-truth membership over the global Actions label universe."],
        ["y_score", "Binary participant selected/not-selected values over the same labels."],
        ["Interpretation", "With binary y_score, row-level ROC-AUC equals balanced accuracy."],
        ["Blank response", "Blank Actions responses are missing, not all-zero predictions, and are excluded from Mean_ROC_AUC."],
        ["Undefined", "Rows whose y_true contains only one class are blank and excluded from Mean_ROC_AUC."],
        ["Action label universe", ", ".join(action_labels)],
    ]
    for row in method_rows:
        method.append(row)
    style_data_header(method[1])
    method.sheet_view.showGridLines = False
    set_widths(method, {"A": 24, "B": 120})
    for row in method.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(analysis_workbook)


def run_complete_analysis(
    temporary_workbook: Path, survey_input: Path, ground_truth_input: Path
):
    """
    Execute the validated analysis stages from raw inputs.

    Stage 1 builds Activity Type and Actions Occurring scoring.
    Stage 2 adds Engagement and Gaze scoring with design-aware frame matching.
    Stage 3 adds Within-subject response-change details and summaries.
    Stage 4 constructs the corrected package-based Between reliability matrices
    directly, without importing the standalone Within/Between helper scripts.
    """
    import build_activity_scoring_workbook as activity_scoring
    import add_engagement_gaze_scoring as engagement_gaze
    import add_within_change_rates as within_change

    # Stage 1: score Activity Type and Actions Occurring from raw survey data.
    survey, participant_map, question_text, activity_gt = (
        activity_scoring.load_inputs(survey_input, ground_truth_input)
    )
    scoring_rows, normalized_gt, scoring_issues = activity_scoring.build_scoring_rows(
        survey, participant_map, question_text, activity_gt
    )
    activity_scoring.write_workbook(
        scoring_rows, normalized_gt, scoring_issues, temporary_workbook
    )

    # Add row-level Actions ROC-AUC and its summary statistics before the other
    # task-specific sheets are appended.
    add_actions_auc(temporary_workbook)

    # Stage 2: add Engagement and Gaze scoring to the same temporary workbook.
    survey, participant_map, question_text, engagement_gt, gaze_gt = (
        engagement_gaze.load_data(survey_input, ground_truth_input)
    )
    engagement_rows, engagement_issues = engagement_gaze.engagement_rows(
        survey, participant_map, question_text, engagement_gt
    )
    gaze_rows, gaze_issues, gaze_mapping = engagement_gaze.gaze_rows(
        survey, participant_map, question_text, gaze_gt
    )
    engagement_gaze.write_outputs(
        engagement_rows, engagement_issues, gaze_rows, gaze_issues,
        gaze_mapping, temporary_workbook,
    )

    # Stage 3: build all Within-subject pair-level and summary change results.
    within_change.add_within_change_sheets(temporary_workbook)

    # Store every Within-subject Engagement comparison, including its ordinal
    # scores, Delta, and direction, so the Table 5 calculations are auditable.
    add_within_engagement_delta_sheet(temporary_workbook)

    # Stage 4: build corrected Between-subject matrices directly in memory.
    ensure_krippendorff_importable()
    records = calculate_between_reliability(temporary_workbook)
    matrices, metadata = reliability_records_to_content(records)
    return matrices, metadata


def alpha_tables_from_metadata(metadata: dict):
    """
    Build 15 Between Alpha summary values and 32 Action-option Alpha values.

    Activity/Engagement/Gaze: use the corresponding matrix metadata alpha.
    Actions: each modality has eight 0/1 matrices; take the arithmetic mean
    after excluding undefined/NaN values.
    """
    direct = {}
    action_values = defaultdict(list)
    action_rows = []
    for key in sorted(metadata):
        meta = metadata[key]
        task = meta["task"]
        modality = meta["modality"]
        alpha = safe_number(meta.get("alpha"))
        if task == "Actions Occurring":
            action_rows.append([
                key, modality, meta.get("action_option_number"),
                meta.get("action_option", ""), alpha,
                int(alpha is not None),
            ])
            if alpha is not None:
                action_values[modality].append(alpha)
        else:
            direct[(task, modality)] = alpha

    summary = {}
    for task in ("Activity Type", "Engagement", "Gaze"):
        for modality in [DISPLAY_MODALITY[m] for m in MODALITIES]:
            summary[(task, modality)] = direct.get((task, modality))
    for modality in [DISPLAY_MODALITY[m] for m in MODALITIES]:
        values = action_values.get(modality, [])
        summary[("Actions Occurring", modality)] = (
            sum(values) / len(values) if values else None
        )
    return summary, action_rows


# ========================================================================
# PART 4. CREATE FINAL_RESULT_CLEAN.XLSX
# ========================================================================


def add_reliability_to_workbook(workbook, matrices, metadata, alpha_summary, action_rows):
    index_ws = workbook.create_sheet("Reliability_Matrix_Index")
    index_ws.append([
        "Matrix_Key", "Excel_Sheet", "Task", "Modality",
        "Action_Option_Number", "Action_Option", "Measurement_Level",
        "Krippendorff_Alpha",
        "M_Raters", "N_Units", "Valid_Cells", "Missing_Cells",
    ])
    style_data_header(index_ws[1])
    index_ws.freeze_panes = "A2"
    index_ws.sheet_view.showGridLines = False

    used_names = set(workbook.sheetnames)
    for key in sorted(matrices):
        matrix = matrices[key]
        meta = metadata[key]
        base_name = "KA_" + clean(meta.get("sheet") or key)
        name = base_name[:31]
        suffix = 1
        while name in used_names:
            tail = f"_{suffix}"
            name = base_name[: 31 - len(tail)] + tail
            suffix += 1
        used_names.add(name)

        valid_cells = int(np.sum(~np.isnan(matrix)))
        missing_cells = int(np.sum(np.isnan(matrix)))
        index_ws.append([
            key, name, meta["task"], meta["modality"],
            meta.get("action_option_number"), meta.get("action_option", ""),
            meta["measurement_level"],
            safe_number(meta.get("alpha")), matrix.shape[0], matrix.shape[1],
            valid_cells, missing_cells,
        ])
        index_ws.cell(index_ws.max_row, 5).number_format = "00"
        index_ws.cell(index_ws.max_row, 8).number_format = "0.000000"

        matrix_ws = workbook.create_sheet(name)
        matrix_ws.append(["Participant / Rater", *meta["clip_ids"]])
        for participant, values in zip(meta["participants"], matrix):
            matrix_ws.append([
                participant,
                *[None if np.isnan(value) else float(value) for value in values],
            ])
        style_data_header(matrix_ws[1])
        matrix_ws.freeze_panes = "B2"
        matrix_ws.sheet_view.showGridLines = False
        matrix_ws.column_dimensions["A"].width = 20
        matrix_ws.row_dimensions[1].height = 48
        for column in range(2, matrix.shape[1] + 2):
            matrix_ws.column_dimensions[get_column_letter(column)].width = 36
        for row in matrix_ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

    set_widths(index_ws, {
        "A": 14, "B": 27, "C": 22, "D": 26, "E": 22, "F": 42,
        "G": 22, "H": 22, "I": 12, "J": 12, "K": 14, "L": 14,
    })

    summary_ws = workbook.create_sheet("Between_Alpha_Summary")
    summary_ws.append(["Modality", *TASKS])
    for modality in [DISPLAY_MODALITY[m] for m in MODALITIES]:
        summary_ws.append([
            modality,
            *[alpha_summary.get((task, modality)) for task in TASKS],
        ])
    style_data_header(summary_ws[1])
    summary_ws.freeze_panes = "B2"
    summary_ws.sheet_view.showGridLines = False
    set_widths(summary_ws, {"A": 26, "B": 20, "C": 22, "D": 20, "E": 16})
    for row in summary_ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = "0.000000"

    action_ws = workbook.create_sheet("Between_Action_Alpha")
    action_ws.append([
        "Matrix_Key", "Modality", "Action_Option_Number", "Action_Option",
        "Krippendorff_Alpha", "Included_In_Modality_Mean",
    ])
    for row in action_rows:
        action_ws.append(row)
    style_data_header(action_ws[1])
    action_ws.freeze_panes = "A2"
    action_ws.sheet_view.showGridLines = False
    set_widths(action_ws, {
        "A": 14, "B": 26, "C": 22, "D": 44, "E": 22, "F": 28,
    })
    for cell in action_ws["C"][1:]:
        cell.number_format = "00"
    for cell in action_ws["E"][1:]:
        cell.number_format = "0.000000"

    method_ws = workbook.create_sheet("Reliability_Method")
    method_rows = [
        ["Item", "Explanation"],
        ["Scope", "Between-subject Krippendorff reliability matrices."],
        ["Rows", "40 participants/raters (P1-P40)."],
        ["Columns", "15 Between stimuli; Gaze uses Clip_ID + Frame_Idx to avoid merging separate frames."],
        ["Missing", "Unassigned modality or blank answer is np.nan in the calculation matrix and blank in Excel."],
        ["Activity Type", "Nominal: Individual=0, Small Group=1, Whole Class=2."],
        ["Actions Occurring", "Eight separate nominal 0/1 matrices per modality; final task value is the mean of valid option-level alphas."],
        ["Engagement", "Ordinal: Low=0, Medium=1, High=2."],
        ["Gaze", "Nominal: ID1=0, ID2=1, ID3=2, ID4=3."],
        ["Source", "Calculated directly from the raw survey and ground truth in this run."],
        ["", ""],
        ["Matrix suffix", "Action option"],
        *[
            [f"{number:02d}", option[0].upper() + option[1:] + "?"]
            for number, option in enumerate(ACTION_OPTIONS, start=1)
        ],
    ]
    for row in method_rows:
        method_ws.append(row)
    style_data_header(method_ws[1])
    style_data_header(method_ws[len(method_rows) - len(ACTION_OPTIONS)])
    method_ws.sheet_view.showGridLines = False
    set_widths(method_ws, {"A": 24, "B": 115})
    for row in method_ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def replace_readme(workbook) -> None:
    """Replace the long build log with a concise scoring-method description."""
    if "README" in workbook.sheetnames:
        position = workbook.sheetnames.index("README")
        del workbook["README"]
        readme = workbook.create_sheet("README", position)
    else:
        readme = workbook.create_sheet("README", 0)

    rows = [
        ["Activity & Actions scoring workbook", ""],
        ["Purpose", "Score 40 participants against the normalized ground truth and summarize four modalities."],
        ["Score", "TP is the number of ground-truth correct labels selected by the participant. Incorrect unselected options do not earn points."],
        ["Proportion correct", "TP divided by the number of ground-truth correct labels. This is equivalent to recall for positive ground-truth labels."],
        ["F1", "2 × precision × recall / (precision + recall). Mean F1 is the arithmetic mean across participant-video response rows."],
        ["AUC", "For each participant-video Actions row, the ground-truth and selected/not-selected status of every action label form y_true and binary y_score. sklearn.metrics.roc_auc_score calculates ROC-AUC. With binary predictions, ROC-AUC equals (sensitivity + specificity) / 2. Undefined and blank rows are excluded from the mean."],
        ["Blank response", "The participant has no recorded response for that participant-video item; the value is treated as missing, not zero."],
        ["0 response", "The participant was assigned to the item and explicitly selected the response option coded as 0."],
        ["Modality summary", "Proportion Correct is pooled TP divided by pooled ground-truth label count. Mean F1 and Mean AUC are arithmetic means of valid participant-video row values."],
    ]
    for row in rows:
        readme.append(row)
    readme.merge_cells("A1:B1")
    readme["A1"].font = Font(size=15, bold=True, color="FFFFFF")
    readme["A1"].fill = PatternFill("solid", fgColor=NAVY)
    readme["A1"].alignment = Alignment(horizontal="center")
    for row in readme.iter_rows(min_row=2):
        row[0].font = Font(bold=True, color=NAVY)
        row[0].fill = PatternFill("solid", fgColor=PALE_BLUE)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    set_widths(readme, {"A": 28, "B": 120})
    readme.sheet_view.showGridLines = False


def build_final_result_clean(
    analysis_workbook: Path, matrices, metadata, alpha_summary, action_rows
) -> None:
    """
    Build the complete Excel-compatible result workbook.

    Some Excel versions report "repaired unreadable content" after openpyxl
    performs a second save of a workbook containing Table, Chart, or
    Conditional Formatting relationship files. This is not numeric or matrix
    corruption.

    The final workbook retains:
      - every worksheet, cell value, formula, number format, and basic style;
      - all 43 reliability matrices and all summary/index/method sheets.

    To maximize compatibility, the final workbook removes:
      - Excel Table objects (the underlying cell data remains intact);
      - Conditional Formatting rules;
      - Chart/Image drawing objects.

    This removes the OOXML relationship features most likely to cause
    cross-version Excel compatibility warnings.
    """
    from openpyxl.formatting.formatting import ConditionalFormattingList
    from openpyxl.worksheet.table import TableList

    workbook = load_workbook(analysis_workbook)
    for worksheet in workbook.worksheets:
        worksheet._tables = TableList()
        worksheet.conditional_formatting = ConditionalFormattingList()
        worksheet._charts = []
        worksheet._images = []
    replace_readme(workbook)
    add_reliability_to_workbook(
        workbook, matrices, metadata, alpha_summary, action_rows
    )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    FINAL_RESULT_CLEAN.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(FINAL_RESULT_CLEAN)


# ========================================================================
# PART 5. STORE WITHIN-SUBJECT ENGAGEMENT DELTA AUDIT RECORDS
# ========================================================================


def add_within_engagement_delta_sheet(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    engagement_rows = [
        row for row in rows_as_dicts(workbook["Engagement_Scoring"])
        if row["Design"] == "Within"
    ]
    lookup = {
        (row["Participant"], row["Clip_ID"], row["Modality"]): row
        for row in engagement_rows
    }
    units = sorted({
        (row["Participant"], row["Clip_ID"]) for row in engagement_rows
    })
    transitions = [
        ("Rendered (no sound) → Rendered", "Rendered without sound", "Rendered with sound"),
        ("Rendered → Blurred", "Rendered with sound", "Blurred"),
        ("Blurred → Raw", "Blurred", "Raw"),
        ("Overall (no sound → raw)", "Rendered without sound", "Raw"),
    ]

    sheet_name = "Within_Engagement_Delta"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append([
        "Participant", "Clip_ID", "Transition", "From_Modality", "To_Modality",
        "Before_Response", "After_Response", "Before_Score", "After_Score",
        "Delta", "Direction", "Valid_Pair",
    ])
    style_data_header(worksheet[1])

    for transition, from_modality, to_modality in transitions:
        for participant, clip_id in units:
            before = lookup.get(
                (participant, clip_id, from_modality), {}
            ).get("Response")
            after = lookup.get(
                (participant, clip_id, to_modality), {}
            ).get("Response")
            valid = before in ENGAGEMENT_SCALE and after in ENGAGEMENT_SCALE
            before_score = ENGAGEMENT_SCALE[before] if before in ENGAGEMENT_SCALE else None
            after_score = ENGAGEMENT_SCALE[after] if after in ENGAGEMENT_SCALE else None
            delta = after_score - before_score if valid else None
            direction = (
                "Up" if valid and delta > 0
                else "Down" if valid and delta < 0
                else "Same" if valid
                else "Missing"
            )
            worksheet.append([
                participant, clip_id, transition, from_modality, to_modality,
                before, after, before_score, after_score, delta, direction,
                int(valid),
            ])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = [12, 38, 34, 28, 25, 18, 18, 14, 14, 10, 12, 12]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    workbook.save(workbook_path)


# ========================================================================
# PART 6. EXTRACT TABLE 2: BETWEEN ACCURACY FROM SCORING SHEETS
# ========================================================================


def mean(values):
    values = [float(value) for value in values if value is not None]
    return sum(values) / len(values) if values else None


def performance_table(source_values_workbook):
    response_rows = rows_as_dicts(source_values_workbook["Response_Scoring"])
    engagement_rows = rows_as_dicts(source_values_workbook["Engagement_Scoring"])
    gaze_rows = rows_as_dicts(source_values_workbook["Gaze_Scoring"])
    table = []
    for modality in MODALITIES:
        display = DISPLAY_MODALITY[modality]
        activity = [
            row for row in response_rows
            if row["Design"] == "Between"
            and row["Question_Type"] == "Activities"
            and row["Modality"] == modality
        ]
        actions = [
            row for row in response_rows
            if row["Design"] == "Between"
            and row["Question_Type"] == "Actions"
            and row["Modality"] == modality
        ]
        engagement = [
            row for row in engagement_rows
            if row["Design"] == "Between" and row["Modality"] == modality
        ]
        gaze_modality = "Rendered" if modality == "Rendered with sound" else modality
        gaze = [
            row for row in gaze_rows
            if row["Design"] == "Between" and row["Modality"] == gaze_modality
        ] if modality != "Rendered without sound" else []

        table.append([
            display,
            mean(row["Proportion_Correct"] for row in activity),
            mean(row["F1"] for row in actions),
            mean(row.get("ROC_AUC") for row in actions),
            mean(row["Correct"] for row in engagement),
            mean(row["Absolute_Distance"] for row in engagement),
            mean(row["Correct"] for row in gaze),
        ])
    return table


# ========================================================================
# PART 7. EXTRACT TABLE 4: WITHIN OPINION CHANGE RATES
# ========================================================================


def change_rate_table(source_values_workbook):
    rows = rows_as_dicts(source_values_workbook["Within_Change_Summary"])
    keep_tasks = {"Activity type", "Actions occurring", "Gaze"}
    output = []
    for row in rows:
        if row["Task"] not in keep_tasks:
            continue
        changed = float(row["Change_Rate"]) * 100
        output.append([
            row["Task"], row["Transition"], changed, 100 - changed,
            int(row["Valid_Pairs"]),
        ])
    return output


# ========================================================================
# PART 8. EXTRACT TABLE 5: ENGAGEMENT CHANGE DIRECTION
# ========================================================================


def engagement_direction_table(source_values_workbook):
    rows = rows_as_dicts(source_values_workbook["Within_Engagement_Delta"])
    transitions = [
        "Rendered (no sound) → Rendered",
        "Rendered → Blurred",
        "Blurred → Raw",
        "Overall (no sound → raw)",
    ]
    output = []
    for label in transitions:
        deltas = [
            float(row["Delta"])
            for row in rows
            if row["Transition"] == label
            and row["Valid_Pair"] == 1
            and row["Delta"] is not None
        ]
        n = len(deltas)
        output.append([
            label,
            100 * sum(delta > 0 for delta in deltas) / n,
            100 * sum(delta < 0 for delta in deltas) / n,
            100 * sum(delta == 0 for delta in deltas) / n,
            sum(deltas) / n,
            n,
        ])
    return output


# ========================================================================
# PART 9. CREATE THE FOUR PAPER-READY TABLES IN FINAL_TABLE.XLSX
# ========================================================================


def setup_paper_sheet(worksheet, title: str, max_column: int):
    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    title_cell = worksheet.cell(1, 1, title)
    title_cell.font = Font(name="Times New Roman", size=14, bold=False)
    title_cell.alignment = Alignment(horizontal="center")
    worksheet.row_dimensions[1].height = 28


def paper_header(worksheet, row_number: int, max_column: int):
    for cell in worksheet[row_number][:max_column]:
        cell.font = Font(name="Times New Roman", size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=BLACK_MEDIUM, bottom=BLACK_THIN)
    worksheet.row_dimensions[row_number].height = 34


def paper_body(worksheet, min_row: int, max_row: int, max_column: int):
    for row in worksheet.iter_rows(
        min_row=min_row, max_row=max_row, min_col=1, max_col=max_column
    ):
        for cell in row:
            cell.font = Font(name="Times New Roman", size=12)
            cell.alignment = Alignment(
                horizontal="left" if cell.column <= 2 else "center",
                vertical="center",
            )
    for cell in worksheet[max_row][:max_column]:
        cell.border = Border(bottom=BLACK_MEDIUM)


def bold_best(worksheet, data_rows, columns, higher_is_better):
    """Bold the best valid value in each requested column; exclude None."""
    for column in columns:
        candidates = [
            worksheet.cell(row, column).value
            for row in data_rows
            if isinstance(worksheet.cell(row, column).value, (int, float))
        ]
        if not candidates:
            continue
        best = max(candidates) if higher_is_better[column] else min(candidates)
        for row in data_rows:
            cell = worksheet.cell(row, column)
            if isinstance(cell.value, (int, float)) and abs(cell.value - best) < 1e-12:
                cell.font = copy(cell.font)
                cell.font = Font(name="Times New Roman", size=12, bold=True)


def add_performance_sheet(workbook, rows):
    ws = workbook.create_sheet("Table2_Accuracy")
    setup_paper_sheet(ws, "Table 2: Accuracy by modality", 7)
    ws.append([
        "Modality", "Activity Type\nProp. Correct", "Actions Occ.\nMean F1",
        "Actions Occ.\nMean ROC-AUC", "Engagement\nProp. Correct",
        "Engagement\nAbs. Dist.", "Gaze\nProp. Correct",
    ])
    paper_header(ws, 2, 7)
    for row in rows:
        ws.append(row)
    paper_body(ws, 3, 6, 7)
    for row in range(3, 7):
        for column in range(2, 8):
            ws.cell(row, column).number_format = "0.000"
        if ws.cell(row, 7).value is None:
            ws.cell(row, 7, "—")
            ws.cell(row, 7).alignment = Alignment(horizontal="center")
    bold_best(
        ws, range(3, 7), [2, 3, 4, 5, 6, 7],
        {2: True, 3: True, 4: True, 5: True, 6: False, 7: True},
    )
    set_widths(ws, {
        "A": 28, "B": 22, "C": 20, "D": 23, "E": 22, "F": 19, "G": 20,
    })
    ws.freeze_panes = "B3"


def add_alpha_sheet(workbook, alpha_summary):
    ws = workbook.create_sheet("Table3_Krippendorff")
    setup_paper_sheet(ws, "Table 3: Between-subject Krippendorff's alpha", 5)
    ws.append(["Modality", "Activity Type", "Actions Occurring", "Engagement", "Gaze"])
    paper_header(ws, 2, 5)
    for modality in [DISPLAY_MODALITY[m] for m in MODALITIES]:
        ws.append([modality, *[alpha_summary.get((task, modality)) for task in TASKS]])
    paper_body(ws, 3, 6, 5)
    for row in range(3, 7):
        for column in range(2, 6):
            ws.cell(row, column).number_format = "0.00"
        if ws.cell(row, 5).value is None:
            ws.cell(row, 5, "—")
    bold_best(ws, range(3, 7), [2, 3, 4, 5], {2: True, 3: True, 4: True, 5: True})
    set_widths(ws, {"A": 28, "B": 20, "C": 23, "D": 19, "E": 15})
    ws.freeze_panes = "B3"


def add_change_sheet(workbook, rows):
    ws = workbook.create_sheet("Table4_Change_Rates")
    setup_paper_sheet(ws, "Table 4: All-modalities opinion change rates per transition", 5)
    ws.append(["Construct", "Transition", "Changed (%)", "Same (%)", "N pairs"])
    paper_header(ws, 2, 5)
    for row in rows:
        ws.append(row)
    last = 2 + len(rows)
    paper_body(ws, 3, last, 5)
    for row in range(3, last + 1):
        ws.cell(row, 3).number_format = "0.0"
        ws.cell(row, 4).number_format = "0.0"
        ws.cell(row, 5).number_format = "0"
    # Add separators at construct boundaries to match a paper-table layout.
    for row in range(4, last + 1):
        if ws.cell(row, 1).value != ws.cell(row - 1, 1).value:
            for cell in ws[row][:5]:
                cell.border = Border(top=BLACK_THIN)
    set_widths(ws, {"A": 24, "B": 46, "C": 18, "D": 16, "E": 14})
    ws.freeze_panes = "C3"


def add_engagement_direction_sheet(workbook, rows):
    ws = workbook.create_sheet("Table5_Engagement_Direction")
    setup_paper_sheet(ws, "Table 5: Engagement opinion change direction per transition", 6)
    ws.append(["Transition", "Up (%)", "Down (%)", "Same (%)", "Mean Δ", "N pairs"])
    paper_header(ws, 2, 6)
    for row in rows:
        ws.append(row)
    paper_body(ws, 3, 6, 6)
    for row in range(3, 7):
        for column in range(2, 5):
            ws.cell(row, column).number_format = "0.0"
        ws.cell(row, 5).number_format = "+0.00;-0.00;0.00"
        ws.cell(row, 6).number_format = "0"
    for cell in ws[6][:6]:
        cell.border = Border(top=BLACK_THIN, bottom=BLACK_MEDIUM)
    set_widths(ws, {"A": 42, "B": 15, "C": 15, "D": 15, "E": 15, "F": 14})
    ws.freeze_panes = "B3"


def build_final_table(alpha_summary):
    """Read scoring data from Final_Result_Clean and generate four table sheets."""
    source_values = load_workbook(
        FINAL_RESULT_CLEAN, read_only=True, data_only=True
    )
    performance = performance_table(source_values)
    change_rates = change_rate_table(source_values)
    engagement_direction = engagement_direction_table(source_values)

    workbook = Workbook()
    workbook.remove(workbook.active)
    add_performance_sheet(workbook, performance)
    add_alpha_sheet(workbook, alpha_summary)
    add_change_sheet(workbook, change_rates)
    add_engagement_direction_sheet(workbook, engagement_direction)
    workbook.save(FINAL_TABLE)


# ========================================================================
# PART 9. FINAL QUALITY ASSURANCE
# ========================================================================

def verify_outputs(original_sheet_names):
    result = load_workbook(FINAL_RESULT_CLEAN, read_only=True, data_only=True)
    missing_original = set(original_sheet_names) - set(result.sheetnames)
    if missing_original:
        raise AssertionError(f"Final_Result lost original sheets: {sorted(missing_original)}")
    matrix_sheets = [name for name in result.sheetnames if name.startswith("KA_")]
    if len(matrix_sheets) != 43:
        raise AssertionError(f"Expected 43 KA_ matrix sheets, found {len(matrix_sheets)}")
    index = result["Reliability_Matrix_Index"]
    if index.max_row != 44:
        raise AssertionError(
            f"Reliability_Matrix_Index should contain 43 records, "
            f"got {index.max_row - 1}"
        )

    tables = load_workbook(FINAL_TABLE, read_only=True, data_only=True)
    expected = {
        "Table2_Accuracy": (6, 7),
        "Table3_Krippendorff": (6, 5),
        "Table4_Change_Rates": (10, 5),
        "Table5_Engagement_Direction": (6, 6),
    }
    if set(tables.sheetnames) != set(expected):
        raise AssertionError(f"Unexpected Final_Table sheets: {tables.sheetnames}")
    for sheet, (rows, columns) in expected.items():
        ws = tables[sheet]
        if (ws.max_row, ws.max_column) != (rows, columns):
            raise AssertionError(
                f"{sheet}: expected {rows}×{columns}, got {ws.max_row}×{ws.max_column}"
            )


def main():
    require_inputs()
    with tempfile.TemporaryDirectory(prefix="rasika_final_") as temporary_dir:
        temporary_path = Path(temporary_dir)
        survey_input, ground_truth_input = prepare_analysis_inputs(temporary_path)
        temporary_workbook = temporary_path / "analysis_workbook.xlsx"
        matrices, metadata = run_complete_analysis(
            temporary_workbook, survey_input, ground_truth_input
        )
        original = load_workbook(temporary_workbook, read_only=True)
        original_sheet_names = list(original.sheetnames)
        original.close()

        alpha_summary, action_rows = alpha_tables_from_metadata(metadata)
        build_final_result_clean(temporary_workbook, matrices, metadata, alpha_summary, action_rows)
        build_final_table(alpha_summary)
        verify_outputs(original_sheet_names)

    print(f"Created: {FINAL_RESULT_CLEAN}")
    print(f"Created: {FINAL_TABLE}")
    print(f"Analysis sheets generated: {len(original_sheet_names)}")
    print("Reliability matrices exported: 43")
    print("Final paper tables exported: 4")

if __name__ == "__main__":
    main()
