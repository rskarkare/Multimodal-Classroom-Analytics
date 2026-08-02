import re
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference


MODALITY_LABELS = {
    "raw": "Raw",
    "blurred": "Blurred",
    "anonymized": "Rendered with sound",
    "anonymized_nosound": "Rendered without sound",
}

WITHIN_FIELDS = {
    "Rendered without sound": ("Q20", "Q21"),
    "Rendered with sound": ("Q26", "Q29"),
    "Blurred": ("Q25", "Q27"),
    "Raw": ("Q24", "Q28"),
}

TYPE_LABELS = {
    "activity_type": "Activities",
    "actions_occurring": "Actions",
}


def clean_text(value):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return str(value).strip()


def split_labels(value):
    text = clean_text(value)
    if not text:
        return []
    return [x.strip() for x in text.split(",") if x.strip()]


def norm_label(value):
    s = clean_text(value).lower()
    s = s.replace("’", "'").replace("?", "")
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"^is teacher ", "is a teacher ", s)
    s = re.sub(r"^is student ", "is a student ", s)
    return s


def canonical_set(value, choice_map):
    output, unknown = set(), []
    for raw in split_labels(value):
        key = norm_label(raw)
        if key in choice_map:
            output.add(choice_map[key])
        else:
            unknown.append(raw)
    return output, unknown


def style_sheet(ws, freeze="A2", autofilter=True):
    navy = "17365D"
    pale = "DCE6F1"
    thin = Side(style="thin", color="D9E1F2")
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = freeze
    if autofilter and ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    for cell in ws[2] if ws.max_row >= 2 else []:
        cell.fill = PatternFill("solid", fgColor=pale)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_table(ws, name):
    if ws.max_row < 2:
        return
    tab = Table(displayName=name, ref=ws.dimensions)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def load_inputs(survey_path, ground_truth_path):
    raw = pd.read_excel(
        survey_path, sheet_name="Sheet 1 - VideoAnonymization_Pi", header=None
    )
    headers = [
        clean_text(x) if clean_text(x) else f"BLANK_{i}"
        for i, x in enumerate(raw.iloc[0].tolist())
    ]
    survey = raw.iloc[2:].copy()
    survey.columns = headers
    survey = survey.reset_index(drop=True)

    participant_map = pd.read_excel(survey_path, sheet_name="Sheet1")
    p_lookup = dict(zip(participant_map["ResponseId"], participant_map["P_label"]))

    metadata = pd.read_excel(survey_path, sheet_name="Sheet2", header=None)
    technical = metadata.iloc[0].tolist()
    questions = metadata.iloc[1].tolist()
    question_text = {
        clean_text(t): clean_text(q)
        for t, q in zip(technical, questions)
        if clean_text(t)
    }

    gt = pd.read_excel(ground_truth_path, sheet_name="Activity", header=1)
    gt = gt[gt["study_design"].isin(["between", "within"])].copy()
    return survey, p_lookup, question_text, gt


def get_clip_from_question(question_text, field):
    text = question_text.get(field, "")
    if " - Which " not in text:
        raise ValueError(f"Cannot extract clip ID from question metadata for {field}: {text}")
    return text.split(" - Which ", 1)[0].strip()


def build_scoring_rows(survey, p_lookup, question_text, gt):
    gt_lookup = {}
    normalized_gt = []
    for _, row in gt.iterrows():
        choices = split_labels(row["answer_choices"])
        gt_set = {norm_label(x) for x in split_labels(row["gt_answer"])}
        key = (clean_text(row["study_design"]), clean_text(row["clip_id"]), clean_text(row["question_type"]))
        gt_lookup[key] = {
            "gt_raw": clean_text(row["gt_answer"]),
            "choices_raw": ", ".join(choices),
            "gt_set": gt_set,
        }
        normalized_gt.append([
            key[0], key[1], key[2], TYPE_LABELS[key[2]],
            ", ".join(sorted(gt_set)), ", ".join(choices), len(choices)
        ])

    rows = []
    problems = []
    for _, respondent in survey.iterrows():
        rid = clean_text(respondent["ResponseId"])
        participant = p_lookup.get(rid, "")
        if not participant:
            problems.append(f"No P label for {rid}")

        assigned_modality = {}
        for slot in range(1, 16):
            clip = clean_text(respondent[f"Clip_{slot}_ID"])
            raw_mod = clean_text(respondent[f"Clip_{slot}_Modality"])
            assigned_modality[clip] = MODALITY_LABELS.get(raw_mod, raw_mod)

        for q_index in range(1, 16):
            activity_field = f"{q_index}_Q17"
            action_field = f"{q_index}_Q13"
            clip = get_clip_from_question(question_text, activity_field)
            modality = assigned_modality.get(clip, "")
            if not modality:
                problems.append(f"{participant}: no between modality assignment for {clip}")
            for qtype, field in [("activity_type", activity_field), ("actions_occurring", action_field)]:
                key = ("between", clip, qtype)
                if key not in gt_lookup:
                    problems.append(f"Missing GT: {key}")
                    continue
                rows.append(score_one(
                    participant, rid, "Between", q_index, clip, modality,
                    qtype, field, respondent.get(field), gt_lookup[key]
                ))

        for q_index in range(1, 5):
            reference_field = f"{q_index}_Q20"
            clip = get_clip_from_question(question_text, reference_field)
            for modality, (activity_q, action_q) in WITHIN_FIELDS.items():
                for qtype, suffix in [("activity_type", activity_q), ("actions_occurring", action_q)]:
                    field = f"{q_index}_{suffix}"
                    key = ("within", clip, qtype)
                    if key not in gt_lookup:
                        problems.append(f"Missing GT: {key}")
                        continue
                    rows.append(score_one(
                        participant, rid, "Within", q_index, clip, modality,
                        qtype, field, respondent.get(field), gt_lookup[key]
                    ))

    return rows, normalized_gt, problems


def score_one(participant, rid, design, slot, clip, modality, qtype, field, response, gt):
    response_set = {norm_label(x) for x in split_labels(response)}
    truth = gt["gt_set"]
    tp = len(response_set & truth)
    fp = len(response_set - truth)
    fn = len(truth - response_set)
    tn = ""
    correct = tp
    total = len(truth)
    precision = tp / (tp + fp) if tp + fp else (1.0 if not truth else 0.0)
    recall = tp / (tp + fn) if tp + fn else 1.0
    f1 = 2 * tp / (2 * tp + fp + fn) if 2 * tp + fp + fn else 1.0
    return [
        participant, rid, design, slot, clip, modality, TYPE_LABELS[qtype], field,
        clean_text(response), ", ".join(sorted(response_set)),
        gt["gt_raw"], ", ".join(sorted(truth)), gt["choices_raw"],
        tp, tn, fp, fn, correct, total, correct / total if total else None,
        precision, recall, f1, int(response_set == truth),
    ]


DETAIL_HEADERS = [
    "Participant", "ResponseId", "Design", "Question_Group", "Clip_ID", "Modality",
    "Question_Type", "Source_Field", "Response_Raw", "Response_Normalized",
    "Ground_Truth_Raw", "Ground_Truth_Normalized", "Answer_Choices",
    "TP", "TN_Not_Scored", "FP", "FN", "Correct_GT_Labels", "Total_GT_Labels",
    "Proportion_Correct", "Precision", "Recall", "F1", "Exact_Match",
]


def write_workbook(rows, normalized_gt, problems, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    readme = [
        ["Activity & Actions scoring workbook", ""],
        ["Purpose", "Score 40 participants against the supplied ground-truth workbook and summarize four modalities."],
        ["Score", "TP: the number of ground-truth correct labels selected by the participant."],
        ["Proportion correct", "TP / number of ground-truth correct labels."],
        ["F1", "2TP / (2TP + FP + FN). This measures agreement on positively selected labels."],
        ["Modality summary", "Proportion Correct is pooled TP / pooled ground-truth label count; Mean F1 is the mean of row-level F1."],
        ["Label note", "Survey 'Is a student raising hand?' and GT 'Is a student sitting?' remain distinct. Raising hand is an FP when selected; a missed GT student-sitting label is an FN."],
        ["Rows scored", len(rows)],
        ["Validation issues", len(problems)],
    ]
    for row in readme:
        ws.append(row)
    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="17365D")
    ws["A1"].alignment = Alignment(horizontal="left")
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).font = Font(bold=True, color="17365D")
        ws.cell(r, 1).fill = PatternFill("solid", fgColor="DCE6F1")
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110
    ws.sheet_view.showGridLines = False

    gt_ws = wb.create_sheet("Ground_Truth_Normalized")
    gt_ws.append(["Study_Design", "Clip_ID", "Question_Type_Code", "Question_Type",
                  "GT_Normalized", "Answer_Choices", "Total_Options"])
    for row in normalized_gt:
        gt_ws.append(row)
    style_sheet(gt_ws)
    set_widths(gt_ws, {"A": 15, "B": 42, "C": 22, "D": 16, "E": 80, "F": 110, "G": 14})
    add_table(gt_ws, "GroundTruthTable")

    detail = wb.create_sheet("Response_Scoring")
    detail.append(DETAIL_HEADERS)
    for row in rows:
        detail.append(row)
    style_sheet(detail)
    set_widths(detail, {
        "A": 12, "B": 22, "C": 11, "D": 15, "E": 42, "F": 25, "G": 16,
        "H": 16, "I": 70, "J": 70, "K": 70, "L": 70, "M": 105,
        "N": 8, "O": 8, "P": 8, "Q": 8, "R": 15, "S": 14,
        "T": 18, "U": 12, "V": 12, "W": 12, "X": 13,
    })
    for col in ["T", "U", "V", "W"]:
        for cell in detail[col][1:]:
            cell.number_format = "0.000"
    add_table(detail, "ResponseScoringTable")
    detail.conditional_formatting.add(
        f"T2:T{detail.max_row}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                       mid_type="num", mid_value=0.5, mid_color="FFEB84",
                       end_type="num", end_value=1, end_color="63BE7B"),
    )
    detail.conditional_formatting.add(
        f"W2:W{detail.max_row}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                       mid_type="num", mid_value=0.5, mid_color="FFEB84",
                       end_type="num", end_value=1, end_color="63BE7B"),
    )

    summary = wb.create_sheet("Modality_Summary")
    summary_headers = [
        "Design", "Question_Type", "Modality", "Scored_Rows",
        "Correct_GT_Label_Points", "Available_GT_Label_Points", "Proportion_Correct",
        "Mean_F1", "Mean_Precision", "Mean_Recall", "Exact_Match_Rate",
    ]
    summary.append(summary_headers)
    modalities = ["Raw", "Blurred", "Rendered with sound", "Rendered without sound"]
    combinations = []
    for design in ["All", "Between", "Within"]:
        for qtype in ["All", "Activities", "Actions"]:
            for modality in modalities:
                combinations.append((design, qtype, modality))
    detail_end = len(rows) + 1
    for design, qtype, modality in combinations:
        summary.append([design, qtype, modality])
        r = summary.max_row
        criteria = [f"'Response_Scoring'!$F$2:$F${detail_end}", f"$C{r}"]
        if design != "All":
            criteria += [f"'Response_Scoring'!$C$2:$C${detail_end}", f"$A{r}"]
        if qtype != "All":
            criteria += [f"'Response_Scoring'!$G$2:$G${detail_end}", f"$B{r}"]
        criteria_text = ",".join(criteria)
        summary.cell(r, 4, f'=COUNTIFS({criteria_text})')
        summary.cell(r, 5, f'=SUMIFS(\'Response_Scoring\'!$R$2:$R${detail_end},{criteria_text})')
        summary.cell(r, 6, f'=SUMIFS(\'Response_Scoring\'!$S$2:$S${detail_end},{criteria_text})')
        summary.cell(r, 7, f'=IFERROR(E{r}/F{r},0)')
        summary.cell(r, 8, f'=IFERROR(AVERAGEIFS(\'Response_Scoring\'!$W$2:$W${detail_end},{criteria_text}),0)')
        summary.cell(r, 9, f'=IFERROR(AVERAGEIFS(\'Response_Scoring\'!$U$2:$U${detail_end},{criteria_text}),0)')
        summary.cell(r, 10, f'=IFERROR(AVERAGEIFS(\'Response_Scoring\'!$V$2:$V${detail_end},{criteria_text}),0)')
        summary.cell(r, 11, f'=IFERROR(AVERAGEIFS(\'Response_Scoring\'!$X$2:$X${detail_end},{criteria_text}),0)')
    style_sheet(summary)
    set_widths(summary, {
        "A": 12, "B": 17, "C": 27, "D": 13, "E": 22, "F": 23,
        "G": 20, "H": 14, "I": 16, "J": 14, "K": 18,
    })
    for col in range(7, 12):
        for r in range(2, summary.max_row + 1):
            summary.cell(r, col).number_format = "0.000"
    add_table(summary, "ModalitySummaryTable")
    summary.conditional_formatting.add(
        f"G2:H{summary.max_row}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                       mid_type="num", mid_value=0.5, mid_color="FFEB84",
                       end_type="num", end_value=1, end_color="63BE7B"),
    )

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Overall performance by modality"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Modality"
    chart.height = 8
    chart.width = 16
    data = Reference(summary, min_col=7, max_col=8, min_row=1, max_row=5)
    cats = Reference(summary, min_col=3, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    summary.add_chart(chart, "M2")

    participant = wb.create_sheet("Participant_Summary")
    participant.append([
        "Participant", "ResponseId", "Modality", "Scored_Rows",
        "Correct_GT_Label_Points", "Available_GT_Label_Points", "Proportion_Correct", "Mean_F1"
    ])
    participants = sorted({(r[0], r[1]) for r in rows}, key=lambda x: int(re.sub(r"\D", "", x[0]) or 0))
    for p_label, rid in participants:
        for modality in modalities:
            participant.append([p_label, rid, modality])
            rr = participant.max_row
            crit = (
                f"'Response_Scoring'!$A$2:$A${detail_end},$A{rr},"
                f"'Response_Scoring'!$F$2:$F${detail_end},$C{rr}"
            )
            participant.cell(rr, 4, f"=COUNTIFS({crit})")
            participant.cell(rr, 5, f"=SUMIFS('Response_Scoring'!$R$2:$R${detail_end},{crit})")
            participant.cell(rr, 6, f"=SUMIFS('Response_Scoring'!$S$2:$S${detail_end},{crit})")
            participant.cell(rr, 7, f"=IFERROR(E{rr}/F{rr},0)")
            participant.cell(rr, 8, f"=IFERROR(AVERAGEIFS('Response_Scoring'!$W$2:$W${detail_end},{crit}),0)")
    style_sheet(participant)
    set_widths(participant, {"A": 12, "B": 22, "C": 27, "D": 13, "E": 22, "F": 23, "G": 20, "H": 14})
    for col in ["G", "H"]:
        for cell in participant[col][1:]:
            cell.number_format = "0.000"
    add_table(participant, "ParticipantSummaryTable")

    validation = wb.create_sheet("Validation")
    validation.append(["Check", "Result"])
    validation_rows = [
        ["Participants", len({r[0] for r in rows})],
        ["Unique ResponseIds", len({r[1] for r in rows})],
        ["Between Activities rows", sum(1 for r in rows if r[2] == "Between" and r[6] == "Activities")],
        ["Between Actions rows", sum(1 for r in rows if r[2] == "Between" and r[6] == "Actions")],
        ["Between total rows", sum(1 for r in rows if r[2] == "Between")],
        ["Within Activities rows", sum(1 for r in rows if r[2] == "Within" and r[6] == "Activities")],
        ["Within Actions rows", sum(1 for r in rows if r[2] == "Within" and r[6] == "Actions")],
        ["Within total rows", sum(1 for r in rows if r[2] == "Within")],
        ["Total scoring rows", len(rows)],
        ["Unmatched records", len(problems)],
        ["Label-handling note", "Survey 'raising hand' and GT 'student sitting' are kept distinct; this is not an unmatched record."],
    ]
    for row in validation_rows:
        validation.append(row)
    for problem in problems:
        validation.append(["Issue", problem])
    style_sheet(validation)
    set_widths(validation, {"A": 32, "B": 100})

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
