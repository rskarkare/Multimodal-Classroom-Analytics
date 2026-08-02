import math
import itertools
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo


VIDEO_MODALITIES = {
    "raw": "Raw",
    "blurred": "Blurred",
    "anonymized": "Rendered with sound",
    "anonymized_nosound": "Rendered without sound",
}
GAZE_MODALITIES = {
    "raw": "Raw",
    "blurred": "Blurred",
    "anonymized": "Rendered",
}
ENGAGE_WITHIN = {
    "Q36": "Rendered without sound",
    "Q37": "Rendered with sound",
    "Q38": "Blurred",
    "Q39": "Raw",
}
GAZE_WITHIN = {0: "Rendered", 1: "Blurred", 2: "Raw"}
ENGAGEMENT_SCALE = {"Low": 0, "Medium": 1, "High": 2}


def clean(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return ""
    return str(v).strip()


def style_sheet(ws, widths):
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_table(ws, name):
    table = Table(displayName=name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)


def get_clip(question_text, field):
    text = question_text[field]
    return text.split(" - ", 1)[0].strip()


def load_data(survey_path, ground_truth_path):
    raw = pd.read_excel(
        survey_path, sheet_name="Sheet 1 - VideoAnonymization_Pi", header=None
    )
    headers = [clean(x) if clean(x) else f"BLANK_{i}" for i, x in enumerate(raw.iloc[0])]
    survey = raw.iloc[2:].copy()
    survey.columns = headers
    survey = survey.reset_index(drop=True)

    pmap = pd.read_excel(survey_path, sheet_name="Sheet1")
    p_lookup = dict(zip(pmap["ResponseId"], pmap["P_label"]))

    meta = pd.read_excel(survey_path, sheet_name="Sheet2", header=None)
    question_text = {
        clean(t): clean(q)
        for t, q in zip(meta.iloc[0], meta.iloc[1])
        if clean(t)
    }
    engage_gt = pd.read_excel(ground_truth_path, sheet_name="Engagement", header=1)
    gaze_gt = pd.read_excel(ground_truth_path, sheet_name="Gaze", header=1)
    return survey, p_lookup, question_text, engage_gt, gaze_gt


def engagement_rows(survey, p_lookup, question_text, gt):
    lookup = {
        (clean(r.study_design), clean(r.clip_id)): clean(r.gt_answer)
        for _, r in gt.iterrows()
    }
    rows, issues = [], []
    for _, person in survey.iterrows():
        rid = clean(person.ResponseId)
        p = p_lookup[rid]
        assigned = {}
        for slot in range(1, 16):
            clip = clean(person[f"Engage_{slot}_ID"])
            assigned[clip] = VIDEO_MODALITIES[clean(person[f"Engage_{slot}_Modality"])]

        for q in range(1, 16):
            field = f"{q}_Q33"
            clip = get_clip(question_text, field)
            gt_answer = lookup.get(("between", clip), "")
            response = clean(person[field])
            if not gt_answer:
                issues.append(f"Missing Engagement GT: between/{clip}")
            rows.append(engagement_score(p, rid, "Between", q, clip, assigned.get(clip, ""), field, response, gt_answer))

        for q in range(1, 7):
            clip = get_clip(question_text, f"{q}_Q36")
            gt_answer = lookup.get(("within", clip), "")
            if not gt_answer:
                issues.append(f"Missing Engagement GT: within/{clip}")
            for suffix, modality in ENGAGE_WITHIN.items():
                field = f"{q}_{suffix}"
                rows.append(engagement_score(
                    p, rid, "Within", q, clip, modality, field, clean(person[field]), gt_answer
                ))
    return rows, sorted(set(issues))


def engagement_score(p, rid, design, group, clip, modality, field, response, gt):
    pred_num = ENGAGEMENT_SCALE.get(response)
    gt_num = ENGAGEMENT_SCALE.get(gt)
    correct = int(response == gt) if response and gt else 0
    abs_dist = abs(pred_num - gt_num) if pred_num is not None and gt_num is not None else None
    return [p, rid, design, group, clip, modality, field, response, gt, pred_num, gt_num, correct, abs_dist]


def infer_gaze_between_group_pairs(survey, question_text):
    all_pairs_by_clip = {}
    for _, person in survey.iterrows():
        for slot in range(1, 16):
            pair = (clean(person[f"Frame_{slot}_ID"]), int(person[f"Frame_{slot}_Idx"]))
            all_pairs_by_clip.setdefault(pair[0], set()).add(pair)

    modality_index = {"anonymized": 0, "blurred": 1, "raw": 2}
    groups_by_clip = {}
    for group in range(1, 16):
        base = (group - 1) * 3 + 1
        groups_by_clip.setdefault(get_clip(question_text, f"{base}_Q53"), []).append(group)

    mapping = {}
    for clip, groups in groups_by_clip.items():
        candidates = sorted(all_pairs_by_clip.get(clip, set()))
        score_matrix = {}
        for group in groups:
            base = (group - 1) * 3 + 1
            for pair in candidates:
                score = 0
                for _, person in survey.iterrows():
                    assigned_mod = None
                    for slot in range(1, 16):
                        if (clean(person[f"Frame_{slot}_ID"]), int(person[f"Frame_{slot}_Idx"])) == pair:
                            assigned_mod = clean(person[f"Frame_{slot}_Modality"])
                            break
                    observed = [i for i in range(3) if clean(person[f"{base + i}_Q53"])]
                    if len(observed) == 1 and modality_index.get(assigned_mod) == observed[0]:
                        score += 1
                score_matrix[(group, pair)] = score
        best = max(
            itertools.permutations(candidates, len(groups)),
            key=lambda perm: sum(score_matrix[(group, pair)] for group, pair in zip(groups, perm))
        )
        for group, pair in zip(groups, best):
            mapping[group] = pair
    return mapping


def gaze_rows(survey, p_lookup, question_text, gt):
    gt_lookup = {
        (clean(r.study_design), clean(r.clip_id), int(r.frame_idx)): clean(r.gt_answer)
        for _, r in gt.iterrows()
    }
    group_pair = infer_gaze_between_group_pairs(survey, question_text)
    issues = []
    if len(set(group_pair.values())) != 15:
        issues.append("Between gaze group-to-frame mapping is not one-to-one.")

    rows = []
    for _, person in survey.iterrows():
        rid = clean(person.ResponseId)
        p = p_lookup[rid]
        assigned = {}
        for slot in range(1, 16):
            pair = (clean(person[f"Frame_{slot}_ID"]), int(person[f"Frame_{slot}_Idx"]))
            assigned[pair] = GAZE_MODALITIES[clean(person[f"Frame_{slot}_Modality"])]

        for group in range(1, 16):
            pair = group_pair[group]
            base = (group - 1) * 3 + 1
            answered = [(base + i, clean(person[f"{base + i}_Q53"])) for i in range(3)]
            answered = [(n, v) for n, v in answered if v]
            field = f"{answered[0][0]}_Q53" if answered else f"{base}_Q53"
            response = answered[0][1] if answered else ""
            if len(answered) > 1:
                issues.append(f"{p}: expected one Between gaze response for group {group}, found {len(answered)}")
            gt_answer = gt_lookup.get(("between", pair[0], pair[1]), "")
            if not gt_answer:
                issues.append(f"Missing Gaze GT: between/{pair}")
            rows.append([
                p, rid, "Between", group, pair[0], pair[1], assigned.get(pair, ""),
                field, response, gt_answer, int(response == gt_answer) if response and gt_answer else 0
            ])

        within_pairs = []
        for slot in range(1, 13, 3):
            pair = (clean(person[f"Within_Frame_{slot}_ID"]), int(person[f"Within_Frame_{slot}_Idx"]))
            within_pairs.append(pair)
        for group in range(1, 5):
            pair = within_pairs[group - 1]
            for offset in range(3):
                field_no = (group - 1) * 3 + offset + 1
                field = f"{field_no}_Q56"
                response = clean(person[field])
                gt_answer = gt_lookup.get(("within", pair[0], pair[1]), "")
                if not gt_answer:
                    issues.append(f"Missing Gaze GT: within/{pair}")
                rows.append([
                    p, rid, "Within", group, pair[0], pair[1], GAZE_WITHIN[offset],
                    field, response, gt_answer, int(response == gt_answer) if response and gt_answer else 0
                ])
    return rows, sorted(set(issues)), group_pair


def replace_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def write_outputs(
    eng_rows, eng_issues, gaze_rows_data, gaze_issues, gaze_mapping,
    output_path,
):
    wb = load_workbook(output_path)

    eng = replace_sheet(wb, "Engagement_Scoring")
    eng.append([
        "Participant", "ResponseId", "Design", "Question_Group", "Clip_ID", "Modality",
        "Source_Field", "Response", "Ground_Truth", "Response_Scale_0_2",
        "GT_Scale_0_2", "Correct", "Absolute_Distance"
    ])
    for row in eng_rows:
        eng.append(row)
    style_sheet(eng, {
        "A": 12, "B": 22, "C": 11, "D": 15, "E": 20, "F": 27, "G": 15,
        "H": 13, "I": 15, "J": 20, "K": 15, "L": 11, "M": 20
    })
    add_table(eng, "EngagementScoringTable")
    eng.conditional_formatting.add(
        f"M2:M{eng.max_row}",
        ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                       mid_type="num", mid_value=1, mid_color="FFEB84",
                       end_type="num", end_value=2, end_color="F8696B")
    )

    es = replace_sheet(wb, "Engagement_Summary")
    es.append(["Design", "Modality", "Scored_Rows", "Correct", "Prop_Correct", "Mean_Abs_Dist"])
    video_mods = ["Raw", "Blurred", "Rendered with sound", "Rendered without sound"]
    eng_end = len(eng_rows) + 1
    for design in ["All", "Between", "Within"]:
        for mod in video_mods:
            es.append([design, mod])
            r = es.max_row
            criteria = [f"'Engagement_Scoring'!$F$2:$F${eng_end}", f"$B{r}"]
            if design != "All":
                criteria += [f"'Engagement_Scoring'!$C$2:$C${eng_end}", f"$A{r}"]
            c = ",".join(criteria)
            es.cell(r, 3, f"=COUNTIFS({c})")
            es.cell(r, 4, f"=SUMIFS('Engagement_Scoring'!$L$2:$L${eng_end},{c})")
            es.cell(r, 5, f"=IFERROR(D{r}/C{r},0)")
            es.cell(r, 6, f"=IFERROR(AVERAGEIFS('Engagement_Scoring'!$M$2:$M${eng_end},{c}),0)")
    style_sheet(es, {"A": 12, "B": 27, "C": 14, "D": 12, "E": 16, "F": 18})
    for col in ["E", "F"]:
        for cell in es[col][1:]:
            cell.number_format = "0.000"
    add_table(es, "EngagementSummaryTable")

    gaze = replace_sheet(wb, "Gaze_Scoring")
    gaze.append([
        "Participant", "ResponseId", "Design", "Question_Group", "Clip_ID", "Frame_Idx",
        "Modality", "Source_Field", "Response", "Ground_Truth", "Correct"
    ])
    for row in gaze_rows_data:
        gaze.append(row)
    style_sheet(gaze, {
        "A": 12, "B": 22, "C": 11, "D": 15, "E": 42, "F": 12,
        "G": 15, "H": 15, "I": 12, "J": 14, "K": 11
    })
    add_table(gaze, "GazeScoringTable")

    gs = replace_sheet(wb, "Gaze_Summary")
    gs.append(["Design", "Modality", "Scored_Rows", "Correct", "Prop_Correct"])
    gaze_end = len(gaze_rows_data) + 1
    combos = [
        ("All", "Raw"), ("All", "Blurred"), ("All", "Rendered"),
        ("Between", "Raw"), ("Between", "Blurred"), ("Between", "Rendered"),
        ("Within", "Raw"), ("Within", "Blurred"), ("Within", "Rendered"),
    ]
    for design, mod in combos:
        gs.append([design, mod])
        r = gs.max_row
        criteria = [f"'Gaze_Scoring'!$G$2:$G${gaze_end}", f"$B{r}"]
        if design != "All":
            criteria += [f"'Gaze_Scoring'!$C$2:$C${gaze_end}", f"$A{r}"]
        c = ",".join(criteria)
        gs.cell(r, 3, f"=COUNTIFS({c})")
        gs.cell(r, 4, f"=SUMIFS('Gaze_Scoring'!$K$2:$K${gaze_end},{c})")
        gs.cell(r, 5, f"=IFERROR(D{r}/C{r},0)")
    style_sheet(gs, {"A": 12, "B": 18, "C": 14, "D": 12, "E": 16})
    for cell in gs["E"][1:]:
        cell.number_format = "0.000"
    add_table(gs, "GazeSummaryTable")

    val = replace_sheet(wb, "Engagement_Gaze_Validation")
    val.append(["Check", "Result"])
    entries = [
        ["Engagement participants", len({r[0] for r in eng_rows})],
        ["Engagement Between rows", sum(r[2] == "Between" for r in eng_rows)],
        ["Engagement Within rows", sum(r[2] == "Within" for r in eng_rows)],
        ["Engagement total rows", len(eng_rows)],
        ["Engagement blank responses", sum(not r[7] for r in eng_rows)],
        ["Engagement unmatched records", len(eng_issues)],
        ["Gaze participants", len({r[0] for r in gaze_rows_data})],
        ["Gaze Between rows", sum(r[2] == "Between" for r in gaze_rows_data)],
        ["Gaze Within rows", sum(r[2] == "Within" for r in gaze_rows_data)],
        ["Gaze total rows", len(gaze_rows_data)],
        ["Gaze blank responses", sum(not r[8] for r in gaze_rows_data)],
        ["Gaze unmatched records", len(gaze_issues)],
        ["Engagement distance scale", "Low=0, Medium=1, High=2; absolute distance range 0–2"],
        ["Gaze modality note", "Still frames use Rendered, Blurred, Raw; sound is not applicable."],
    ]
    for row in entries:
        val.append(row)
    for issue in eng_issues:
        val.append(["Engagement issue", issue])
    for issue in gaze_issues:
        val.append(["Gaze issue", issue])
    val.append(["Gaze group mapping", "; ".join(f"{g}:{p[0]}@{p[1]}" for g, p in gaze_mapping.items())])
    style_sheet(val, {"A": 34, "B": 140})

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(output_path)
