from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule


VIDEO_TRANSITIONS = [
    ("Rendered without sound", "Rendered with sound", "Rendered (no sound) → Rendered"),
    ("Rendered with sound", "Blurred", "Rendered → Blurred"),
    ("Blurred", "Raw", "Blurred → Raw"),
]
GAZE_TRANSITIONS = [
    ("Rendered", "Blurred", "Rendered → Blurred"),
    ("Blurred", "Raw", "Blurred → Raw"),
]


def clean(value):
    return "" if value is None else str(value).strip()


def rows_as_dicts(ws):
    headers = [c.value for c in ws[1]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]


def style_sheet(ws, widths):
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_table(ws, name):
    table = Table(displayName=name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)


def replace_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def make_pairs(rows, task, transitions, value_col, unit_cols):
    selected = [r for r in rows if r["Design"] == "Within"]
    lookup = {}
    for r in selected:
        key = tuple(r[c] for c in ["Participant", *unit_cols])
        lookup[(key, r["Modality"])] = r

    output = []
    units = sorted({tuple(r[c] for c in ["Participant", *unit_cols]) for r in selected})
    for from_mod, to_mod, transition in transitions:
        for key in units:
            before = lookup.get((key, from_mod))
            after = lookup.get((key, to_mod))
            before_value = clean(before.get(value_col)) if before else ""
            after_value = clean(after.get(value_col)) if after else ""
            valid = int(bool(before_value) and bool(after_value))
            changed = int(before_value != after_value) if valid else None
            participant = key[0]
            clip = key[1]
            extra = key[2] if len(key) > 2 else ""
            output.append([
                participant, task, transition, from_mod, to_mod, clip, extra,
                before_value, after_value, valid, changed,
            ])
    return output


def add_within_change_sheets(output_path):
    wb = load_workbook(output_path)
    response = rows_as_dicts(wb["Response_Scoring"])
    engagement = rows_as_dicts(wb["Engagement_Scoring"])
    gaze = rows_as_dicts(wb["Gaze_Scoring"])

    activity_rows = [
        r for r in response
        if r["Question_Type"] == "Activities"
    ]
    action_rows = [
        r for r in response
        if r["Question_Type"] == "Actions"
    ]
    details = []
    details += make_pairs(
        activity_rows, "Activity type", VIDEO_TRANSITIONS,
        "Response_Normalized", ["Clip_ID"]
    )
    details += make_pairs(
        action_rows, "Actions occurring", VIDEO_TRANSITIONS,
        "Response_Normalized", ["Clip_ID"]
    )
    details += make_pairs(
        engagement, "Engagement", VIDEO_TRANSITIONS,
        "Response", ["Clip_ID"]
    )
    details += make_pairs(
        gaze, "Gaze", GAZE_TRANSITIONS,
        "Response", ["Clip_ID", "Frame_Idx"]
    )

    detail = replace_sheet(wb, "Within_Change_Detail")
    detail.append([
        "Participant", "Task", "Transition", "From_Modality", "To_Modality",
        "Stimulus_ID", "Frame_Idx", "Before_Response", "After_Response",
        "Valid_Pair", "Changed"
    ])
    for row in details:
        detail.append(row)
    style_sheet(detail, {
        "A": 12, "B": 20, "C": 34, "D": 29, "E": 27, "F": 42, "G": 12,
        "H": 85, "I": 85, "J": 12, "K": 12
    })
    add_table(detail, "WithinChangeDetailTable")

    summary = replace_sheet(wb, "Within_Change_Summary")
    summary.append([
        "Task", "Transition", "Potential_Pairs", "Valid_Pairs",
        "Changed_Pairs", "Unchanged_Pairs", "Missing_Pairs", "Change_Rate"
    ])
    order = [
        ("Activity type", [x[2] for x in VIDEO_TRANSITIONS]),
        ("Actions occurring", [x[2] for x in VIDEO_TRANSITIONS]),
        ("Engagement", [x[2] for x in VIDEO_TRANSITIONS]),
        ("Gaze", [x[2] for x in GAZE_TRANSITIONS]),
    ]
    for task, transitions in order:
        for transition in transitions:
            subset = [r for r in details if r[1] == task and r[2] == transition]
            valid = [r for r in subset if r[9] == 1]
            changed = sum(r[10] == 1 for r in valid)
            summary.append([
                task, transition, len(subset), len(valid), changed,
                len(valid) - changed, len(subset) - len(valid),
                changed / len(valid) if valid else None,
            ])
    style_sheet(summary, {
        "A": 20, "B": 34, "C": 17, "D": 14, "E": 16,
        "F": 18, "G": 16, "H": 15
    })
    for cell in summary["H"][1:]:
        cell.number_format = "0.0%"
    add_table(summary, "WithinChangeSummaryTable")
    summary.conditional_formatting.add(
        f"H2:H{summary.max_row}",
        ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                       mid_type="num", mid_value=0.5, mid_color="FFEB84",
                       end_type="num", end_value=1, end_color="F8696B")
    )

    notes = replace_sheet(wb, "Within_Change_Notes")
    note_rows = [
        ["Within-subject change-rate methodology", ""],
        ["Definition", "Change Rate = changed valid pairs / all valid pairs."],
        ["Pairing", "The before and after responses must come from the same participant and the same clip/frame."],
        ["Activity type", "Changed=1 when the normalized activity response differs between adjacent modalities."],
        ["Actions occurring", "Changed=1 when the complete selected-action set differs in any way; ordering does not matter."],
        ["Engagement", "Changed=1 when Low/Medium/High differs."],
        ["Gaze", "Changed=1 when selected person ID differs. Static frames have no no-sound condition."],
        ["Missing response", "If either side is blank, the pair is marked missing and excluded from the change-rate denominator."],
        ["Corrected mapping", "Q20/Q21=Rendered without sound; Q26/Q29=Rendered with sound; Q25/Q27=Blurred; Q24/Q28=Raw."],
        ["Important", "Change Rate measures response instability, not whether the change moves toward or away from ground truth."],
    ]
    for row in note_rows:
        notes.append(row)
    notes.merge_cells("A1:B1")
    notes["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    notes["A1"].fill = PatternFill("solid", fgColor="17365D")
    for r in range(2, notes.max_row + 1):
        notes.cell(r, 1).font = Font(bold=True, color="17365D")
        notes.cell(r, 1).fill = PatternFill("solid", fgColor="DCE6F1")
        notes.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    notes.column_dimensions["A"].width = 25
    notes.column_dimensions["B"].width = 115
    notes.sheet_view.showGridLines = False

    wb.save(output_path)
