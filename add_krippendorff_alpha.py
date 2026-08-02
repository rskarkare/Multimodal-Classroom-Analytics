from pathlib import Path
from collections import Counter, defaultdict

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule


ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "outputs" / "activity_actions_scoring_40_participants.xlsx"

def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def split_set(value):
    text = clean(value)
    return {x.strip() for x in text.split(",") if x.strip()} if text else set()


def krippendorff_alpha(units, level):
    """
    units: list of lists; each inner list contains valid ratings for one unit.
    Implements Krippendorff's coincidence-matrix alpha for nominal or ordinal data.
    """
    usable = [list(x) for x in units if len(x) >= 2]
    categories = sorted({v for unit in usable for v in unit})
    if not usable or len(categories) < 2:
        return {
            "alpha": None, "Do": None, "De": None, "units": len(usable),
            "ratings": sum(len(x) for x in usable), "categories": categories,
            "min_raters": min((len(x) for x in usable), default=0),
            "max_raters": max((len(x) for x in usable), default=0),
        }

    observed = defaultdict(float)
    marginals = Counter()
    for ratings in usable:
        counts = Counter(ratings)
        m = len(ratings)
        for c, nc in counts.items():
            marginals[c] += nc
            for k, nk in counts.items():
                observed[(c, k)] += nc * (nk - (1 if c == k else 0)) / (m - 1)

    n = sum(marginals.values())

    def distance(c, k):
        if c == k:
            return 0.0
        if level == "nominal":
            return 1.0
        lo, hi = sorted((c, k))
        ordered = sorted(categories)
        between = [x for x in ordered if lo <= x <= hi]
        span = sum(marginals[x] for x in between) - (marginals[c] + marginals[k]) / 2
        return span * span

    do = sum(value * distance(c, k) for (c, k), value in observed.items()) / n
    de_num = 0.0
    for c in categories:
        for k in categories:
            expected = marginals[c] * (marginals[k] - (1 if c == k else 0)) / (n - 1)
            de_num += expected * distance(c, k)
    de = de_num / n
    alpha = 1 - do / de if de else None
    return {
        "alpha": alpha, "Do": do, "De": de, "units": len(usable), "ratings": n,
        "categories": categories,
        "min_raters": min(len(x) for x in usable),
        "max_raters": max(len(x) for x in usable),
    }


def rows_as_dicts(ws):
    headers = [cell.value for cell in ws[1]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]


def calculate_corrected(input_workbook=None):
    """Compatibility entry point for the canonical corrected calculation."""
    from build_final_results import calculate_between_reliability

    return calculate_between_reliability(input_workbook or OUTPUT)


def activity_action_alpha(rows, question_type, modality):
    selected = [
        r for r in rows
        if r["Design"] == "Between"
        and r["Question_Type"] == question_type
        and r["Modality"] == modality
        and clean(r["Response_Normalized"])
    ]
    all_task_rows = [
        r for r in rows
        if r["Design"] == "Between"
        and r["Question_Type"] == question_type
        and clean(r["Response_Normalized"])
    ]
    labels = {
        label
        for row in all_task_rows
        for label in split_set(row["Response_Normalized"])
    }
    labels = sorted(labels)
    units = []
    for clip in sorted({r["Clip_ID"] for r in selected}):
        clip_rows = [r for r in selected if r["Clip_ID"] == clip]
        for label in labels:
            units.append([
                1 if label in split_set(r["Response_Normalized"]) else 0
                for r in clip_rows
            ])
    result = krippendorff_alpha(units, "nominal")
    result["label_count"] = len(labels)
    result["stimulus_count"] = len({r["Clip_ID"] for r in selected})
    result["response_rows"] = len(selected)
    return result


def engagement_alpha(rows, modality):
    selected = [
        r for r in rows
        if r["Design"] == "Between" and r["Modality"] == modality
        and r["Response_Scale_0_2"] is not None
    ]
    grouped = defaultdict(list)
    for r in selected:
        grouped[r["Clip_ID"]].append(int(r["Response_Scale_0_2"]))
    result = krippendorff_alpha(list(grouped.values()), "ordinal")
    result["label_count"] = 3
    result["stimulus_count"] = len(grouped)
    result["response_rows"] = len(selected)
    return result


def gaze_alpha(rows, modality):
    selected = [
        r for r in rows
        if r["Design"] == "Between" and r["Modality"] == modality and clean(r["Response"])
    ]
    grouped = defaultdict(list)
    for r in selected:
        grouped[(r["Clip_ID"], r["Frame_Idx"])].append(clean(r["Response"]))
    result = krippendorff_alpha(list(grouped.values()), "nominal")
    result["label_count"] = len({clean(r["Response"]) for r in selected})
    result["stimulus_count"] = len(grouped)
    result["response_rows"] = len(selected)
    return result


def style_sheet(ws, widths):
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_table(ws, name):
    table = Table(displayName=name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)


def replace_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def main():
    wb = load_workbook(OUTPUT)
    response_rows = rows_as_dicts(wb["Response_Scoring"])
    engagement_rows = rows_as_dicts(wb["Engagement_Scoring"])
    gaze_rows = rows_as_dicts(wb["Gaze_Scoring"])

    results = []
    video_modalities = ["Raw", "Blurred", "Rendered with sound", "Rendered without sound"]
    for task in ["Activities", "Actions"]:
        for modality in video_modalities:
            result = activity_action_alpha(response_rows, task, modality)
            results.append((task, modality, "Nominal (binary label expansion)", result))
    for modality in video_modalities:
        result = engagement_alpha(engagement_rows, modality)
        results.append(("Engagement", modality, "Ordinal (Low=0, Medium=1, High=2)", result))
    for modality in ["Raw", "Blurred", "Rendered"]:
        result = gaze_alpha(gaze_rows, modality)
        results.append(("Gaze", modality, "Nominal", result))

    ws = replace_sheet(wb, "Krippendorff_Alpha")
    ws.append([
        "Design", "Task", "Modality", "Measurement_Level", "Alpha",
        "Observed_Disagreement_Do", "Expected_Disagreement_De",
        "Stimuli", "Expanded_Units", "Valid_Ratings", "Response_Rows",
        "Min_Raters_Per_Unit", "Max_Raters_Per_Unit", "Label_Categories"
    ])
    for task, modality, level, result in results:
        ws.append([
            "Between", task, modality, level, result["alpha"], result["Do"], result["De"],
            result["stimulus_count"], result["units"], result["ratings"], result["response_rows"],
            result["min_raters"], result["max_raters"], result["label_count"]
        ])
    style_sheet(ws, {
        "A": 12, "B": 16, "C": 27, "D": 39, "E": 12, "F": 25, "G": 25,
        "H": 11, "I": 16, "J": 15, "K": 15, "L": 20, "M": 20, "N": 18
    })
    for col in ["E", "F", "G"]:
        for cell in ws[col][1:]:
            cell.number_format = "0.000"
    add_table(ws, "KrippendorffAlphaTable")
    ws.conditional_formatting.add(
        f"E2:E{ws.max_row}",
        ColorScaleRule(start_type="num", start_value=-1, start_color="F8696B",
                       mid_type="num", mid_value=0.5, mid_color="FFEB84",
                       end_type="num", end_value=1, end_color="63BE7B")
    )

    notes = replace_sheet(wb, "Krippendorff_Notes")
    note_rows = [
        ["Krippendorff’s Alpha methodology", ""],
        ["Scope", "Between-subject/random-anonymization records only."],
        ["Unit and rater", "Stimulus is the reliability unit; participant is the rater. Participants not assigned to a modality are missing by design."],
        ["Activities", "Nominal alpha after binary expansion: each video × each observed activity label is one unit, rated 0/1 by assigned participants."],
        ["Actions", "Nominal alpha after binary expansion: each video × each observed action label is one unit, rated 0/1 by assigned participants."],
        ["Engagement", "Ordinal alpha using Low=0, Medium=1, High=2 and Krippendorff’s ordinal distance based on category marginals."],
        ["Gaze", "Nominal alpha; unit is clip ID + frame index and rating is the selected person ID."],
        ["Missing responses", "Excluded from alpha as missing ratings. Random modality non-assignment is also represented as missing by design."],
        ["Interpretation", "Alpha=1 means perfect agreement; 0 means chance-level agreement; negative values mean disagreement beyond chance."],
        ["Important", "Alpha measures inter-rater agreement, not correctness versus ground truth."],
    ]
    for row in note_rows:
        notes.append(row)
    notes.merge_cells("A1:B1")
    notes["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    notes["A1"].fill = PatternFill("solid", fgColor="17365D")
    for r in range(2, notes.max_row + 1):
        notes.cell(r, 1).font = Font(bold=True, color="17365D")
        notes.cell(r, 1).fill = PatternFill("solid", fgColor="DCE6F1")
        notes.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    notes.column_dimensions["A"].width = 23
    notes.column_dimensions["B"].width = 115
    notes.sheet_view.showGridLines = False

    wb.save(OUTPUT)
    print(f"saved={OUTPUT}")
    for task, modality, _, result in results:
        print(task, modality, result["alpha"], result["response_rows"], result["units"])


if __name__ == "__main__":
    main()
